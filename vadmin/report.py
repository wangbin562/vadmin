# !/usr/bin/python
# -*- coding=utf-8 -*-
"""
views
"""
import os
import time
import json
import openpyxl
from django.conf import settings
from django.db import models
from django.db.models import fields
from openpyxl.styles import Alignment
from django.db.models.query import QuerySet
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

from vadmin import admin_api
from vadmin import widgets
from vadmin import const
from vadmin import common


def export_excel(request, model_admin, queryset=None, path=None):
    if queryset is None:
        queryset = get_export_queryset(request, model_admin)
    w_book = openpyxl.Workbook()
    w_sheet = w_book.active
    export_config = model_admin.get_export_config(request)
    if "fields" in export_config and export_config["fields"]:
        lst_field_name = export_config["fields"]
    else:
        lst_field_name = model_admin.get_list_display(request)

    width_config = export_config.get("width", {})
    for i, field_name in enumerate(lst_field_name):
        if field_name in width_config:
            width = width_config[field_name]
        elif i in width_config:
            width = width_config[i]
        else:
            width = 120

        from utils import conver
        width_excel = conver.px_2_in(width) * 10.0  # 1/10英寸
        col_name = get_column_letter(i + 1)
        w_sheet.column_dimensions[col_name].width = width_excel
        # w_sheet.column_dimensions[col_name].alignment = Alignment(horizontal='left')

    row = []
    dict_title = export_config.get("title", {})
    for field_name in lst_field_name:
        fun = getattr(model_admin, field_name, None) or \
              getattr(model_admin, "list_v_%s" % field_name, None)

        if field_name in dict_title:
            title = dict_title[field_name]
            row.append(str(title).strip())
            continue

        elif callable(fun):
            if getattr(fun, "multiple_col", False):
                lst_title = getattr(model_admin, "%s_short_description" % fun.__name__)(request)
                for title in lst_title:
                    row.append(str(title).strip())
                continue

            elif hasattr(fun, 'short_description'):
                title = getattr(fun, 'short_description')
                row.append(str(title).strip())
                continue
        try:
            db_field = admin_api.get_field(model_admin, field_name)
            title = db_field.verbose_name
        except (BaseException,):
            title = field_name

        row.append(str(title).strip())

    w_sheet.append(row)

    for obj in queryset:
        row = []
        for field_name in lst_field_name:
            fun = getattr(model_admin, field_name, None) or \
                  getattr(model_admin, "export_v_%s" % field_name, None) or \
                  getattr(model_admin, "list_v_%s" % field_name, None)
            if callable(fun):  # 自定义
                if getattr(fun, "multiple_col", False):
                    lst_value = fun(request, obj)
                    for value in lst_value:
                        if isinstance(value, (int, float, str)):
                            row.append(value)
                        else:
                            row.append(str(value))
                else:
                    value = fun(request, obj)
                    if isinstance(value, (int, float, str)):
                        row.append(value)
                    else:
                        row.append(str(value))
            else:
                if "__" in field_name:
                    field1, field2 = field_name.split("__", 1)
                    try:
                        obj_sub = getattr(obj, field1)
                        value = getattr(obj_sub, field2)
                    except (BaseException,):
                        value = ""
                else:
                    try:
                        value = getattr(obj, field_name)
                    except (BaseException,):
                        value = ""

                if isinstance(value, (int, float, str)):
                    row.append(value)
                else:
                    row.append(str(value))

        w_sheet.append(row)

    if path:
        file_url = common.get_download_url(path)
    else:
        path, file_url = get_export_file_name(request, model_admin)

    w_book.save(path)
    return file_url


def format_cell_value(value, db_field=None):
    number_format = numbers.FORMAT_TEXT

    if db_field:
        if hasattr(db_field, "choices") and db_field.choices:
            dict_choices = dict(db_field.choices)
            try:
                if value in dict_choices:
                    value = dict_choices[value]
                elif "," in value:
                    lst_value = []
                    for single in value.split(","):
                        lst_value.append(dict_choices[single])
                    value = ",".join(lst_value)
            except (BaseException,):
                pass

        if isinstance(db_field, fields.DateTimeField):
            value = value.strftime("%Y-%m-%d %H:%M:%S")
            number_format = numbers.FORMAT_DATE_DATETIME
        elif isinstance(db_field, fields.DateField):
            value = value.strftime("%Y-%m-%d")
            number_format = numbers.FORMAT_DATE_YYYYMMDD2
        elif isinstance(db_field, fields.TimeField):
            value = value.strftime("%H:%M:%S")
            number_format = numbers.FORMAT_DATE_TIME8
        elif isinstance(db_field, (fields.BooleanField, fields.NullBooleanField)):
            dict_bool = {True: u"是", False: u"否"}
            value = dict_bool.get(value, value)
        elif isinstance(db_field, ((fields.IntegerField, fields.PositiveIntegerField, fields.AutoField,
                                    fields.PositiveSmallIntegerField, fields.SmallIntegerField,
                                    fields.BigIntegerField, fields.BigAutoField))):
            number_format = numbers.FORMAT_NUMBER
        elif isinstance(db_field, (fields.FloatField, fields.DecimalField)):
            number_format = numbers.FORMAT_NUMBER_00
        elif value is None:
            value = ""
        else:
            value = str(value)

    elif isinstance(value, bool):
        dict_bool = {True: u"是", False: u"否"}
        value = dict_bool.get(value, value)
    elif isinstance(value, int):
        number_format = numbers.FORMAT_NUMBER
    elif isinstance(value, float):
        number_format = numbers.FORMAT_NUMBER_00
    else:
        number_format = numbers.FORMAT_TEXT
        if value is None:
            value = ""
        elif isinstance(value, widgets.Text):
            value = str(value.get("text", ""))
        else:
            value = str(value)

    return number_format, value


def get_export_queryset(request, model_admin):
    content = request.POST.get(const.SUBMIT_WIDGET, "{}")
    widget_data = json.loads(content)

    app_label = model_admin.opts.app_label
    model_name = model_admin.opts.model_name

    if int(widget_data.get(const.WN_SELECT_ALL, 0) or 0):  # 导出全部
        queryset = admin_api.get_filter_queryset(request, model_admin)
        queryset = model_admin.export_filter(request, queryset)
    else:
        # 获取表格的值
        table_name = const.WN_TABLE % (app_label, model_name)
        try:
            lst_object_id = widget_data[table_name]["value"] or []
        except (BaseException,):
            lst_object_id = []

        queryset = model_admin.model.objects.filter(pk__in=lst_object_id)
        queryset = model_admin.export_filter(request, queryset)
        if isinstance(queryset, QuerySet):
            dict_sort = admin_api.parse_sort_para(request, model_admin)
            lst_order = []
            for k, v in dict_sort.items():
                if v == "+":
                    lst_order.append(k)
                else:
                    lst_order.append("-%s" % k)
            if lst_order:
                queryset = queryset.order_by(*lst_order)  # 界面控制排序
            elif model_admin.ordering:
                queryset = queryset.order_by(*model_admin.ordering)  # 使用model_admin配置排序
            elif model_admin.opts.ordering:
                queryset = queryset.order_by(*model_admin.opts.ordering)  # 使用model配置排序

    return queryset


def get_export_file_name(request, model_admin):
    export_config = model_admin.get_export_config(request)
    app_label = model_admin.opts.app_label
    model_name = model_admin.opts.model_name
    # base_path = os.path.join(settings.MEDIA_ROOT, "export")

    if "file_path" in export_config and export_config["file_path"]:
        base_path = export_config["file_path"]
        path = os.path.abspath(base_path)
        if not os.path.exists(path):
            os.makedirs(path)

        if "file_name" in export_config and export_config["file_name"]:
            file_name = export_config["file_name"]
        else:
            file_name = "%s_%s_%s.xlsx" % (app_label, model_name, request.user.id)

        path = os.path.join(path, file_name)
        path = os.path.abspath(path)
        if os.path.exists(path):
            try:
                os.remove(path)
            except (BaseException,):
                pass

        return path, os.path.join(settings.MEDIA_URL, file_name)

    else:
        file_name = "%s_%s_%s.xlsx" % (app_label, model_name, request.user.id)
        file_path = common.get_export_path(file_name)

        return file_path, common.get_download_url(file_path)
