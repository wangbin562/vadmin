# !/usr/bin/python
# -*- coding=utf-8 -*-
"""
v1.2
"""

import os
import sys
import importlib
import collections
import openpyxl
import codecs
import json
import copy
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.styles import colors, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

if sys.version_info[0] == 2:
    from StringIO import StringIO
else:
    from io import StringIO

# from openpyxl.styles import Color
# from openpyxl.styles import PatternFill
# from openpyxl.styles import Font
# from openpyxl.styles import Alignment

BEGIN_KEY = "{%"
END_KEY = "%}"


def pt_2_px(pt):
    return pt * 4 / 3.0


def split_row_col(row_col):
    """
    分解行列
    :param row_col: 'B11' 'AB1' 'AB123'
    :return:
    """
    row = 0
    idx = 0
    for idx, key in enumerate(row_col):
        key_ascii = ord(key.upper())
        if (key_ascii >= 65) and (key_ascii <= 90):
            pass
        else:
            row = row_col[idx:]
            break

    col = column_index_from_string(row_col[0:idx])
    row = int(row)
    return col, row


def none_2_null_str(value):
    if value is None:
        return ""

    return value


class AbstractSyntaxNode:
    def __str__(self):
        return repr(self)

    def __repr__(self):
        return "<" + self.__class__.__name__ + ">"


class TextNode(AbstractSyntaxNode):
    """
    文本结点
    """

    def __init__(self, text, indent=0):
        self.text = text
        self.indent = indent

    def merge_text(self, text):
        self.text += text


class ScriptNode(AbstractSyntaxNode):
    """
    脚本结点
    """

    def __init__(self, text, indent=0):
        self.text = text.rstrip()
        self.indent = indent


class Compiler(object):
    def __init__(self, template_file, sheet_idx=0, max_row=200, max_col=100):
        self.lst_global = []
        self.lst_border = []
        self.lst_fill = []
        self.lst_font = []
        self.lst_alignment = []
        self.work_book = openpyxl.load_workbook(template_file)
        self.work_sheet = self.work_book.worksheets[sheet_idx]
        self.max_row = max_row
        self.max_col = max_col
        self.lst_cell = []
        self.lst_row_height = []
        self.lst_col_width = []
        self.lst_code = []
        self.stream = None
        self.stream_len = 0
        self.current_idx = 0
        self.lst_node = []

    def compile_page(self, **kwargs):
        for key, val in kwargs.items():
            self.lst_global.append(key)

        self.parse_excel()
        return self.generate_code()

    def check_merged_cell(self, row, col):
        """
        检查合并单元格
        :param row:
        :param col:
        :return: b_merged：是否合并 b_range：是否在合并的区域内 row_span：合并的行数 col_span：合并的列数
        """
        b_merged = False
        b_range = False
        row_span = 0
        col_span = 0
        row_col = '%s%s' % (get_column_letter(col), row)
        # for s_range in self.work_sheet.merged_cell_ranges:
        for s_range in self.work_sheet.merged_cells.ranges:
            s_range = str(s_range)
            begin_row_col, end_row_col = s_range.split(":")
            begin_col, begin_row = split_row_col(begin_row_col)
            end_col, end_row = split_row_col(end_row_col)

            if begin_row_col == row_col:  # 开始合并
                b_merged = True
                row_span = end_row - begin_row
                col_span = end_col - begin_col
                break

            elif (row >= begin_row) and (row <= end_row) and (col >= begin_col) and (col <= end_col):
                b_range = True
                break

        return b_merged, b_range, row_span, col_span

    def _make_default(self, key, value, default, is_str=False, is_u=False):
        if isinstance(default, list):
            if value in default:
                return_value = ''
                return return_value

        if value == default:
            return_value = ''
            return return_value

        if is_str:
            if is_u:
                return_value = '%s=u"%s", ' % (key, value)
            else:
                return_value = '%s="%s", ' % (key, value)
        else:
            return_value = '%s=%s, ' % (key, value)

        return return_value

    def _make_color(self, color):
        if color is None:
            s_color = 'None'
        else:
            if color.type == "theme":
                s_color = "Color(theme=%s,tint=%s,type='%s')" % (color.theme, color.tint, color.type)
            elif color.type == "indexed":
                s_color = "Color(indexed=%s,tint=%s,type='%s')" % (color.indexed, color.tint, color.type)
            elif color.type == "rgb":
                s_color = "Color(rgb='%s',tint=%s,type='%s')" % (color.rgb, color.tint, color.type)
            else:
                s_color = "None"
        return s_color

    def _make_side(self, style, color):
        lst_none = [None, 'None']

        style = self._make_default('style', style, lst_none, True)
        color = self._make_default('color', color, lst_none)

        s_side = 'Side(%s%s' % (style, color)
        s_side = '%s)' % s_side.rstrip(",")
        return s_side

    def parse_cell_style(self, cell):
        # 边框颜色
        left_color = self._make_color(cell.border.left.color)
        right_color = self._make_color(cell.border.right.color)
        top_color = self._make_color(cell.border.top.color)
        bottom_color = self._make_color(cell.border.bottom.color)
        # diagonal_color = self._make_color(cell.border.diagonal.color)

        # 边框样式
        left_side = self._make_side(cell.border.left.style, left_color)
        right_side = self._make_side(cell.border.right.style, right_color)
        top_side = self._make_side(cell.border.top.style, top_color)
        bottom_side = self._make_side(cell.border.bottom.style, bottom_color)
        border = "Border(left=%s,right=%s,top=%s,bottom=%s)" % (left_side, right_side, top_side, bottom_side)
        if border not in self.lst_border:
            self.lst_border.append(border)

        fgColor = self._make_color(cell.fill.fgColor)
        # bgColor = self._make_color(cell.fill.bgColor)
        patternType = self._make_default('patternType', cell.fill.patternType, None, True)
        fill = "PatternFill(%sfgColor=%s)" % (patternType, fgColor)
        if fill not in self.lst_fill:
            self.lst_fill.append(fill)

        name = self._make_default('name', cell.font.name, None, True, True)
        # sz = self._make_default('sz', cell.font.sz, None)
        # b = self._make_default('b', cell.font.b, None)
        # i = self._make_default('i', cell.font.i, None)
        charset = self._make_default('charset', cell.font.charset, None)
        # u = self._make_default('u', cell.font.u, None)
        strike = self._make_default('strike', cell.font.strike, None)
        scheme = self._make_default('scheme', cell.font.scheme, None, True)
        family = self._make_default('family', cell.font.family, None)
        size = self._make_default('size', cell.font.size, None)
        bold = self._make_default('bold', cell.font.bold, None)
        italic = self._make_default('italic', cell.font.italic, None)
        underline = self._make_default('underline', cell.font.underline, default=None, is_str=True)
        vertAlign = self._make_default('vertAlign', cell.font.vertAlign, None)
        outline = self._make_default('outline', cell.font.outline, None)
        shadow = self._make_default('shadow', cell.font.shadow, None)
        condense = self._make_default('condense', cell.font.condense, None)
        extend = self._make_default('extend', cell.font.extend, None)
        color = self._make_color(cell.font.color)
        font = "Font(%s%s%s%s%s%s%s%s%s%s%s%s%s%scolor=%s)" % (
            name, charset, strike, scheme, family, size, bold, italic, underline, vertAlign, outline,
            shadow, condense, extend, color
        )
        if font not in self.lst_font:
            self.lst_font.append(font)

        # horizontal = None, vertical = None,
        # textRotation = 0, wrapText = None, shrinkToFit = None, indent = 0, relativeIndent = 0,
        # justifyLastLine = None, readingOrder = 0, text_rotation = None,
        # wrap_text = None, shrink_to_fit = None, mergeCell = None
        horizontal = self._make_default('horizontal', cell.alignment.horizontal, None, True)
        vertical = self._make_default('vertical', cell.alignment.vertical, None, True)
        indent = self._make_default('indent', cell.alignment.indent, 0)
        relativeIndent = self._make_default('relativeIndent', cell.alignment.relativeIndent, 0)
        justifyLastLine = self._make_default('justifyLastLine', cell.alignment.justifyLastLine, None)
        readingOrder = self._make_default('readingOrder', cell.alignment.readingOrder, 0)
        text_rotation = self._make_default('text_rotation', cell.alignment.text_rotation, [None, 0])
        wrap_text = self._make_default('wrap_text', cell.alignment.wrap_text, None)
        shrink_to_fit = self._make_default('shrink_to_fit', cell.alignment.shrink_to_fit, None)
        alignment = "Alignment(%s%s%s%s%s%s%s%s%s)" % (
            horizontal, vertical, indent, relativeIndent,
            justifyLastLine, readingOrder, text_rotation, wrap_text, shrink_to_fit
        )
        if alignment not in self.lst_alignment:
            self.lst_alignment.append(alignment)

        return border, fill, font, alignment

    def calc_indent(self, key=None):
        # if self.def_num > 0 and key == "def":
        #     indent = 8 * self.def_num
        #
        # elif self.def_num > 0:
        #     indent = 4 + 8 * self.def_num
        #
        # elif self.block_name and key == "def":
        #     indent = 4
        #
        # elif self.block_name:
        #     indent = 8
        # else:
        indent = 8

        return indent

    def add_node(self, node):
        # if self.block_name is not None:
        #     self.dict_method[self.block_name].append(node)
        # else:
        self.lst_node.append(node)

    def parse_cell(self, row, col, row_span, col_span):
        """
        解释文件
        :return:
        """
        lst_node = []
        while True:
            begin_idx, end_idx = self.next_node()
            if (begin_idx < 0) or (end_idx < 0):
                if self.current_idx < self.stream_len:
                    begin_idx = self.stream_len
                else:
                    break

            if begin_idx != self.current_idx:
                text = self.stream[self.current_idx:begin_idx]
                text = text.lstrip("\r\n")
                text = text.rstrip()
                if text != "":
                    node = TextNode(text, self.calc_indent())
                    lst_node.append(node)

                if end_idx < 0:
                    break

            # key = self.stream[begin_idx + 2]
            # if key == "#":
            #     text = self.stream[begin_idx + 3:end_idx]
            #     text = text.strip()
            #     lst_text = text.split(" ", 1)
            #     instr = lst_text[0].strip().lower()
            #     if len(lst_text) > 1:
            #         # value = lst_text[1].strip().replace('"', '').replace("'", "")
            #         value = lst_text[1].strip()
            #     else:
            #         value = None
            #     lst_node.extend(self.process_instruction(instr, value, row, col))
            #
            # elif key == "=":
            #     text = self.stream[begin_idx + 3:end_idx]
            #     text = text.strip()
            #     lst_node.append(AssignmentNode(text, self.calc_indent()))
            #
            # elif key == "!" and self.stream[begin_idx + 2: begin_idx + 5] == "!--":
            #     pass
            #
            # else:
            text = self.stream[begin_idx + 2:end_idx]
            text = text.lstrip("\r\n")
            text = text.rstrip("")
            # if text.lower() == "end":
            #     if self.def_num > 0:
            #         self.def_num -= 1
            #
            # elif text[0:3].lower() == "def":
            #     self.def_num += 1
            #     lst_node.append(ScriptNode(text, self.calc_indent("def")))
            #
            # else:
            lst_node.append(ScriptNode(text, self.calc_indent()))

            self.current_idx = end_idx + 2

        return lst_node

    def next_node(self):
        global BEGIN_KEY, END_KEY
        begin_idx = self.stream.find(BEGIN_KEY, self.current_idx)
        end_idx = self.stream.find(END_KEY, self.current_idx + 2)
        return begin_idx, end_idx


class PageCompiler(Compiler):
    def __init__(self, template_file, full=False,
                 bg_color="#ffffff", font_color='#000000',
                 border_width=1, border_color='#000000',
                 font_size=14.67,
                 horizontal="center", vertical="center"):
        """
        :param full: 是否为全量生成（True：生成所有表格边框 False:统一边框）
        :param bg_color: 默认背景颜色
        :param font_color: 默认文字颜色
        :param border_width: 默认边框宽度 1px
        :param border_color: 默认边框颜色 黑色
        :param font_size: 默认文字大小 14.67 对应excel 11号字体
        :param horizontal: 水平居中 默认center 共3个值 left center right
        :param vertical: 垂直居中 默认center 共3个值 top center bottom
        """
        super(PageCompiler, self).__init__(template_file)
        self.block_name = None
        self.dict_method = collections.OrderedDict()
        self.full = full
        self.default_bg_color = bg_color
        self.default_font_color = font_color
        self.default_border_width = border_width
        self.default_border_color = border_color
        self.default_font_size = font_size
        self.default_horizontal = horizontal
        self.default_vertical = vertical
        self.lst_font = []
        self.lst_alignment = []
        self.lst_fill = []

    def _make_color(self, color):
        if color is None:
            s_color = 'None'
        else:
            if color.type == "theme":
                lst_theme = [
                    "FFFFFF", "000000", "EEECE1", "1F497D", "4F81BD", "C0504D", "9BBB59", "8064A2", "4BACC6", "F79646",
                    "F2F2F2", "7F7F7F", "DDD9C3", "C6D9F0", "DBE5F1", "F2DCDB", "EBF1DD", "E5E0EC", "DBEEF3", "FDEADA",
                    "D8D8D8", "595959", "C4BD97", "8DB3E2", "B8CCE4", "E5B9B7", "D7E3BC", "CCC1D9", "B7DDE8", "FBD5B5",
                    "BFBFBF", "3F3F3F", "938953", "548DD4", "95B3D7", "D99694", "C3D69B", "B2A2C7", "92CDDC", "FAC08F",
                    "A5A5A5", "262626", "494429", "17365D", "366092", "953734", "76923C", "5F497A", "31859B", "E36C09",
                    "7F7F7F", "0C0C0C", "1D1B10", "0F243E", "244061", "632423", "4F6128", "3F3151", "205867", "974806",
                ]
                s_color = "'%s'" % lst_theme[color.theme]

            elif color.type == "indexed":
                if (color.indexed > 0) or (color.indexed < 64):
                    s_color = "'%s'" % colors.COLOR_INDEX[color.indexed][2:]
                else:
                    s_color = "None"

            elif color.type == "rgb":
                if color.rgb == "00000000":
                    s_color = "'FFFFFF'"
                else:
                    s_color = "'%s'" % color.rgb[2:]
            else:
                s_color = "None"
        return s_color

    def parse_excel(self):
        """
        解释文件
        :return:
        """
        max_row = self.work_sheet.max_row + 1
        max_col = self.work_sheet.max_column + 1

        if max_row < self.max_row:
            self.max_row = max_row
        if max_col < self.max_col:
            self.max_col = max_col

        lst_row = [None] * self.max_col
        for i in range(self.max_row):
            self.lst_cell.append(copy.copy(lst_row))

        for row in range(1, self.max_row + 1):
            height = self.work_sheet.row_dimensions[row].height
            if height is None:
                height = 13.5

            height = pt_2_px(height)
            self.lst_row_height.append(height)
            for col in range(1, self.max_col + 1):
                if row == 1:
                    col_name = get_column_letter(col)
                    col_width = self.work_sheet.column_dimensions[col_name].width
                    if not self.work_sheet.column_dimensions[col_name].customWidth:
                        col_width = 9
                    self.lst_col_width.append(col_width)

                cell = self.work_sheet.cell(row=row, column=col)

                fgColor = self._make_color(cell.fill.fgColor)
                patternType = self._make_default('patternType', cell.fill.patternType, None, True)
                if fgColor == 'None':
                    fill = "PatternFill(%s)" % patternType
                else:
                    fill = "PatternFill(%sfgColor=%s)" % (patternType, fgColor)
                if fill not in self.lst_fill:
                    self.lst_fill.append(fill)

                name = self._make_default('name', cell.font.name, None, True, True)
                charset = self._make_default('charset', cell.font.charset, None)
                strike = self._make_default('strike', cell.font.strike, None)
                scheme = self._make_default('scheme', cell.font.scheme, None, True)
                family = self._make_default('family', cell.font.family, None)
                size = self._make_default('size', cell.font.size, None)
                bold = self._make_default('bold', cell.font.bold, None)
                italic = self._make_default('italic', cell.font.italic, None)
                underline = self._make_default('underline', cell.font.underline, default=None, is_str=True)
                vertAlign = self._make_default('vertAlign', cell.font.vertAlign, None)
                outline = self._make_default('outline', cell.font.outline, None)
                shadow = self._make_default('shadow', cell.font.shadow, None)
                condense = self._make_default('condense', cell.font.condense, None)
                extend = self._make_default('extend', cell.font.extend, None)
                color = self._make_color(cell.font.color)
                font = "Font(%s%s%s%s%s%s%s%s%s%s%s%s%s%scolor=%s)" % (
                    name, charset, strike, scheme, family, size, bold, italic, underline, vertAlign, outline,
                    shadow, condense, extend, color
                )
                if font not in self.lst_font:
                    self.lst_font.append(font)

                horizontal = self._make_default('horizontal', cell.alignment.horizontal, None, True)
                vertical = cell.alignment.vertical
                if vertical is None and cell.value:
                    vertical = "bottom"
                vertical = self._make_default('vertical', vertical, None, is_str=True)
                indent = self._make_default('indent', cell.alignment.indent, 0)
                relativeIndent = self._make_default('relativeIndent', cell.alignment.relativeIndent, 0)
                justifyLastLine = self._make_default('justifyLastLine', cell.alignment.justifyLastLine, None)
                readingOrder = self._make_default('readingOrder', cell.alignment.readingOrder, 0)
                text_rotation = self._make_default('text_rotation', cell.alignment.text_rotation, [None, 0])
                wrap_text = self._make_default('wrap_text', cell.alignment.wrap_text, None)
                shrink_to_fit = self._make_default('shrink_to_fit', cell.alignment.shrink_to_fit, None)
                alignment = "Alignment(%s%s%s%s%s%s%s%s%s)" % (
                    horizontal, vertical, indent, relativeIndent,
                    justifyLastLine, readingOrder, text_rotation, wrap_text, shrink_to_fit
                )
                if alignment not in self.lst_alignment:
                    self.lst_alignment.append(alignment)

                # 判读是否合并
                b_merged, b_range, row_span, col_span = self.check_merged_cell(row, col)
                self.lst_cell[row - 1][col - 1] = "%s, %s, %s, %s, %s, %s, %s, %s, %s" % (
                    row - 1, col - 1, self.lst_fill.index(fill),
                    self.lst_font.index(font), self.lst_alignment.index(alignment),
                    b_merged, b_range, row_span, col_span)

                if b_range:
                    continue

                self.stream = none_2_null_str(cell.value)
                if cell.data_type in ['n']:
                    self.stream = str(self.stream)
                self.stream_len = len(self.stream)
                self.current_idx = 0

                lst_node = self.parse_cell(row, col, row_span, col_span)
                if lst_node:
                    text = "def write_%s_%s(self, r_cell, _w):" % (row - 1, col - 1)
                    self.add_node(ScriptNode(text, 4))
                    self.lst_node.extend(lst_node)
                    self.add_node(ScriptNode("", 0))

    def generate_code(self):
        """
        生成代码
        """
        lst_code = []
        lst_code.append('# !/usr/bin/python')
        lst_code.append('# -*- coding: utf-8 -*-')
        lst_code.append('"""')
        lst_code.append('code generated by Epage compiler')
        lst_code.append('"""')
        lst_code.append('''import sys
if sys.version_info[0] == 2:
    from StringIO import StringIO
else:
    from io import StringIO
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment''')
        lst_code.append('')
        for s_global in self.lst_global:
            lst_code.append("%s = None" % s_global)
        lst_code.append('')
        lst_code.append('''
class Page(object):
    def __init__(self, out):
        self.out = out
        self._w = self.out.write
        self.dict_add_row = {}
        self.dict_add_col = {}
        self.dict_w_cell = {}
        self.dict_height = {}''')
        text = "        self.lst_fill = []"
        lst_code.append(text)
        for fill in self.lst_fill:
            text = "        self.lst_fill.append(%s)" % fill
            lst_code.append(text)

        text = "        self.lst_font = []"
        lst_code.append(text)
        for font in self.lst_font:
            text = "        self.lst_font.append(%s)" % font
            lst_code.append(text)

        text = "        self.lst_alignment = []"
        lst_code.append(text)
        for alignment in self.lst_alignment:
            text = "        self.lst_alignment.append(%s)" % alignment
            lst_code.append(text)
        text = "        self.lst_cell = []"
        lst_code.append(text)
        for i, lst_row in enumerate(self.lst_cell):
            text = "        self.lst_cell.append([])"
            lst_code.append(text)
            for cell in lst_row:
                text = "        self.lst_cell[%s].append([%s])" % (i, cell)
                lst_code.append(text)
        text = "        self.lst_row_height = %s" % str(self.lst_row_height)
        lst_code.append(text)
        text = "        self.lst_col_width = %s" % str(self.lst_col_width)
        lst_code.append(text)

        lst_code.append('''
    def get_cell_style(self, fill_idx, font_idx, alignment_idx):
        fill = self.lst_fill[fill_idx]
        font = self.lst_font[font_idx]
        alignment = self.lst_alignment[alignment_idx]
        s_style = ""
        if fill.fgColor:
            if fill.fgColor.rgb not in [None, "00FFFFFF"]:
                s_style = 'background-color:#%s;' % fill.fgColor.rgb[2:]

        if font.color:
            if font.color.rgb not in [None, "00000000"]:
                s_style += 'color:#%s;' % font.color.rgb[2:]

        if font.size != 11:
            font_size = font.size * 4 / 3.0
            s_style += "font-size:%spx;" % font_size

        if font.bold:
            s_style += "font-weight:bold;"

        if font.italic:
            s_style += "font-style:italic;"

        if font.underline in ["single", "double"]:
            s_style += "text-decoration:underline;"

        if alignment.horizontal not in ["center", None]:
            s_style += 'text-align:%s;' % alignment.horizontal

        if alignment.vertical not in ["center", None]:
            s_style += 'vertical-align:%s;' % alignment.vertical

        return s_style
        
    def create(self, **kwargs):''')

        for s_global in self.lst_global:
            lst_code.append('        global %s' % s_global)
            lst_code.append('        %s = kwargs["%s"]' % (s_global, s_global))
        lst_code.append('''
        for i, lst_row in enumerate(self.lst_cell):
            w_row = i + self.dict_add_row.get(0, 0)
            self.dict_height[w_row] = i
            for r_cell in lst_row:
                r_row, r_col, fill_idx, font_idx, alignment_idx, b_merged, b_range, row_span, col_span = r_cell

                if b_range:
                    continue

                output = StringIO()
                _w = output.write
                w_row = r_row + self.dict_add_row.get(r_col, 0)
                w_col = r_col + self.dict_add_col.get(r_row, 0)
                write_fun = getattr(self, "write_%s_%s" % (r_row, r_col), None)
                if write_fun is not None:
                    b_loop = write_fun(r_cell, _w)
                    if not b_loop:
                        _w('</td>')

                        output_front = StringIO()
                        _w = output_front.write
                        _w('<td')
                        if r_row == 1:
                            width = round(float(self.lst_col_width[r_col]) / sum(self.lst_col_width) * 100.0, 2)
                            _w(' width="%s%%"' % width)

                        if col_span:
                            _w(' colspan="%s"' % (col_span + 1))

                        if row_span:
                            _w(' rowspan="%s"' % (row_span + 1))

                        s_style = self.get_cell_style(fill_idx, font_idx, alignment_idx)
                        if s_style:
                            _w(' style="%s"' % s_style)
                        _w('>')
                        self.dict_w_cell.setdefault(w_row, {})[w_col] = output_front.getvalue() + output.getvalue()
                else:
                    _w('<td')
                    if r_row == 1:
                        width = round(float(self.lst_col_width[r_col]) / sum(self.lst_col_width) * 100.0, 2)
                        _w(' width="%s%%"' % width)

                    if col_span:
                        _w(' colspan="%s"' % (col_span + 1))

                    if row_span:
                        _w(' rowspan="%s"' % (row_span + 1))

                    s_style = self.get_cell_style(fill_idx, font_idx, alignment_idx)
                    if s_style:
                        _w(' style="%s"' % s_style)
                    _w('>')
                    _w('</td>')
                    self.dict_w_cell.setdefault(w_row, {})[w_col] = output.getvalue()

        _w = self.out.write
        _w(
            '<table border="1" cellspacing="0" cellpadding="0" style="border-collapse:collapse;border:1px solid #000000;text-align:center;vertical-align:center;color:#000000;font-size:14.67px;background-color:#ffffff;word-break:break-all;">')
        _w('\\n')

        max_row = 0
        max_col = 0
        if self.lst_cell:
            if self.dict_add_row:
                max_row = len(self.lst_cell) + max(self.dict_add_row.values())
            else:
                max_row = len(self.lst_cell)

            if self.dict_add_col:
                max_col = len(self.lst_cell[0]) + max(self.dict_add_col.values())
            else:
                max_col = len(self.lst_cell[0])

        for row in range(max_row):
            _w('<tr style="height:%spx;">' % int(self.lst_row_height[self.dict_height.get(row, row)]))
            for col in range(max_col):
                if row in self.dict_w_cell and col in self.dict_w_cell[row]:
                    _w('\\n')
                    _w(self.dict_w_cell[row][col])
            _w('</tr>')
            _w('\\n')
        _w('</table>')
''')

        for node in self.lst_node:
            self._generate_code(node, lst_code)

        lst_code.append('')
        return "\n".join(lst_code)

    def _generate_temp_code(self, lst_node):
        lst_code = []
        for node in lst_node:
            self._generate_code(node, lst_code)

        lst_code.append('')
        return "\n".join(lst_code)

    def _generate_code(self, node, lst_code):
        indent = " " * node.indent
        if isinstance(node, TextNode):
            for text in node.text.split("\n"):
                if "'" in text:
                    line = '%s_w("%s")' % (indent, text)
                else:
                    line = "%s_w('%s')" % (indent, text)
                lst_code.append(line)

        elif isinstance(node, ScriptNode):
            for text in node.text.split("\n"):
                text = text.rstrip()
                text = text.replace("\t", "    ")
                lst_code.append(indent + text)

        else:
            # raise CompilerError("invalid syntax element " + repr(c))
            print(node.text)


class ExcelCompiler(Compiler):
    def __init__(self, template_file, sheet_idx=0, max_row=200, max_col=100):
        super(ExcelCompiler, self).__init__(template_file, sheet_idx, max_row, max_col)

    def parse_excel(self):
        """
        解释文件
        """

        max_row = self.work_sheet.max_row + 1
        max_col = self.work_sheet.max_column + 1
        if max_row < self.max_row:
            self.max_row = max_row
        if max_col < self.max_col:
            self.max_col = max_col

        lst_row = [None] * self.max_col
        for i in range(self.max_row):
            self.lst_cell.append(copy.copy(lst_row))

        for row in range(1, self.max_row + 1):
            self.lst_row_height.append(self.work_sheet.row_dimensions[row].height)
            for col in range(1, self.max_col + 1):
                if row == 1:
                    col_name = get_column_letter(col)
                    col_width = self.work_sheet.column_dimensions[col_name].width
                    if not self.work_sheet.column_dimensions[col_name].customWidth:
                        col_width = 9
                    self.lst_col_width.append(col_width)

                cell = self.work_sheet.cell(row=row, column=col)
                border, fill, font, alignment = self.parse_cell_style(cell)

                # 判读是否合并
                b_merged, b_range, row_span, col_span = self.check_merged_cell(row, col)
                self.lst_cell[row - 1][col - 1] = "%s, %s, %s, %s, %s, %s, %s, %s, %s, %s" % (
                    row, col, self.lst_border.index(border), self.lst_fill.index(fill),
                    self.lst_font.index(font), self.lst_alignment.index(alignment),
                    b_merged, b_range, row_span, col_span)

                if b_range:
                    continue

                self.stream = none_2_null_str(cell.value)
                if cell.data_type in ['n']:
                    self.stream = str(self.stream)
                self.stream_len = len(self.stream)
                self.current_idx = 0
                #
                lst_node = self.parse_cell(row, col, row_span, col_span)
                if lst_node:
                    # node = lst_node[0]
                    text = "def write_%s_%s(self):" % (row, col)
                    self.add_node(ScriptNode(text, 4))
                    for node in lst_node:
                        self.add_node(node)
                    self.add_node(ScriptNode("", 0))
                    self.add_node(ScriptNode("", 0))

    def _generate_code(self, node):
        indent = " " * node.indent
        if isinstance(node, TextNode):
            for text in node.text.split("\n"):
                if "'" in text:
                    line = '%s_w("%s")' % (indent, text)
                else:
                    line = "%s_w('%s')" % (indent, text)
                self.lst_code.append(line)

        # elif isinstance(node, AssignmentNode):
        #     line = "%s_w(str(%s))" % (indent, node.text)
        #     lst_code.append(line)

        elif isinstance(node, ScriptNode):
            for text in node.text.split("\n"):
                text = text.rstrip()
                text = text.replace("\t", "    ")
                self.lst_code.append(indent + text)

        # elif isinstance(node, CommentNode):
        #     lst_code.append('')
        #     lst_code.append(indent + node.text)

        else:
            # raise CompilerError("invalid syntax element " + repr(c))
            print(node.text)

    def generate_import(self):
        self.lst_code.append('import traceback')
        self.lst_code.append('import copy')
        self.lst_code.append('import openpyxl')
        self.lst_code.append('from openpyxl.utils import get_column_letter')
        self.lst_code.append('from openpyxl.utils import column_index_from_string')
        self.lst_code.append('from openpyxl.styles import colors, Color')
        self.lst_code.append('from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment')
        self.lst_code.append('')
        self.lst_code.append('')

    def generate_method(self):
        text = '''
    def _copy_cell_style(self):
        """
        拷贝单元格样式
        """
        if self.row_span or self.col_span:
            border = self.lst_border[self.border_idx]
            for row in range(self.row_span + 1):
                for col in range(self.col_span + 1):
                    cell = self.w_sheet.cell(row=self.w_row + row, column=self.w_col + col)
                    cell.border = copy.deepcopy(border)
        else:  
            border = copy.deepcopy(self.lst_border[self.border_idx])
            self.w_cell.border = border
                    
        fill = copy.deepcopy(self.lst_fill[self.fill_idx])
        font = copy.deepcopy(self.lst_font[self.font_idx])
        alignment = copy.deepcopy(self.lst_alignment[self.alignment_idx])
        
        self.w_cell.fill = fill
        self.w_cell.font = font
        self.w_cell.alignment = alignment


    def _add_cell_down(self):
        """
        向下增加一个单元格
        """
        if self.row_span or self.col_span:
            if self.row_span:
                add_row = self.dict_add_row.get(self.r_col, 0) + self.row_span + 1
                for row in range(self.row_span + 1):
                    row_height = self.lst_row_height[self.r_row + row - 1]
                    self.w_sheet.row_dimensions[self.r_row + add_row + row].height = row_height  # 设置行高
                self.w_row = self.r_row + add_row
                
            if self.col_span:
                add_col = self.dict_add_col.get(self.r_row, 0) + self.col_span + 1
                # for col in range(self.col_span + 1):
                #     col_name = get_column_letter(self.r_col + add_col + col)
                #     self.w_sheet.column_dimensions[col_name].width = self.lst_col_width[self.r_col + col - 1] # 设置列宽
                self.w_col = self.r_col + add_col
                
            s_range = "%s%s:%s%s" % (get_column_letter(self.w_col), self.w_row, get_column_letter(self.w_col + self.col_span),
                                     self.w_row + self.row_span)
            self.w_sheet.merge_cells(s_range)  # 合并单元格
            
            self.w_cell = self.w_sheet.cell(row=self.w_row, column=self.w_col)
            self._copy_cell_style()
            
            if self.r_col in self.dict_add_row:
                self.dict_add_row[self.r_col] += (self.row_span + 1)
            else:
                self.dict_add_row[self.r_col] = self.row_span + 1
                
            # if self.r_row in self.dict_add_col:
            #     self.dict_add_col[self.r_row] += (self.col_span + 1)
            # else:
            #     self.dict_add_col[self.r_row] = self.col_span + 1

        else:
            row_height = self.lst_row_height[self.r_row]
            self.w_row += 1
            if self.r_col in self.dict_add_row:
                self.dict_add_row[self.r_col] += 1
            else:
                self.dict_add_row[self.r_col] = 1
            self.w_sheet.row_dimensions[self.w_row].height = row_height  # 设置行高
            self.w_cell = self.w_sheet.cell(row=self.w_row, column=self.w_col)
            self._copy_cell_style()
        

    def _add_cell_right(self):
        pass


    def _write_style(self, **kwargs):
        """
        写入样式
        """
        if "background_color" in kwargs:
            fill = PatternFill(patternType="solid", fgColor=Color(rgb='FF%s' % kwargs["background_color"][1:],tint=0.0,type='rgb'))
            self.w_cell.fill = fill
            
        if ("font_color" in kwargs) or ("font_size" in kwargs):
            old_font = self.w_cell.font
            if "font_color" in kwargs:
                font = Font(name=old_font.name, size=kwargs.get("font_size", old_font.size), color=Color(rgb='FF%s' % kwargs["font_color"][1:],tint=0.0,type='rgb'))
            else:
                font = Font(name=old_font.name, size=kwargs.get("font_size", old_font.size))
            self.w_cell.font = font
'''
        self.lst_code.append(text)

    def generate_code(self):
        """
        生成代码
        """
        self.lst_code.append('# !/usr/bin/python')
        self.lst_code.append('# -*- coding: utf-8 -*-')
        self.lst_code.append('"""')
        self.lst_code.append('code generated by Epage compiler')
        self.lst_code.append('"""')
        self.generate_import()

        for s_global in self.lst_global:
            self.lst_code.append("%s = None" % s_global)

        self.lst_code.append('_w = None')
        self.lst_code.append('_copy_cell_style = None')
        self.lst_code.append('_add_cell_down = None')
        self.lst_code.append('_add_cell_right = None')
        self.lst_code.append('_write_style = None')
        self.lst_code.append('')

        self.lst_code.append('''
class Page(object):
    def __init__(self, output_file):
        self.w_cell = None
        self.border_idx = None
        self.fill_idx = None
        self.font_idx = None
        self.alignment_idx = None
        self.r_row = None
        self.r_col = None
        self.w_row = None
        self.w_col = None
        self.output_file = output_file
        self.w_book = openpyxl.Workbook()
        self.w_sheet = self.w_book.active
        self.dict_add_row = {}
        self.dict_add_col = {}
        self.lst_w_cell = []
''')
        text = "        self.lst_border = []"
        self.lst_code.append(text)
        for board in self.lst_border:
            text = "        self.lst_border.append(%s)" % board
            self.lst_code.append(text)

        text = "        self.lst_fill = []"
        self.lst_code.append(text)
        for fill in self.lst_fill:
            text = "        self.lst_fill.append(%s)" % fill
            self.lst_code.append(text)

        text = "        self.lst_font = []"
        self.lst_code.append(text)
        for font in self.lst_font:
            text = "        self.lst_font.append(%s)" % font
            self.lst_code.append(text)

        text = "        self.lst_alignment = []"
        self.lst_code.append(text)
        for alignment in self.lst_alignment:
            text = "        self.lst_alignment.append(%s)" % alignment
            self.lst_code.append(text)

        text = "        self.lst_cell = []"
        self.lst_code.append(text)
        for i, lst_row in enumerate(self.lst_cell):
            text = "        self.lst_cell.append([])"
            self.lst_code.append(text)
            for cell in lst_row:
                text = "        self.lst_cell[%s].append([%s])" % (i, cell)
                self.lst_code.append(text)
        text = "        self.lst_row_height = %s" % str(self.lst_row_height)
        self.lst_code.append(text)
        text = "        self.lst_col_width = %s" % str(self.lst_col_width)
        self.lst_code.append(text)
        self.lst_code.append("")
        self.generate_method()

        self.lst_code.append('    def create(self, **kwargs):')
        for s_global in self.lst_global:
            self.lst_code.append('        global %s' % s_global)
            self.lst_code.append('        %s = kwargs["%s"]' % (s_global, s_global))

        self.lst_code.append('        global _w')
        self.lst_code.append('        _w = self._w')
        self.lst_code.append('        global _copy_cell_style')
        self.lst_code.append('        _copy_cell_style = self._copy_cell_style')
        self.lst_code.append('        global _add_cell_down')
        self.lst_code.append('        _add_cell_down = self._add_cell_down')
        self.lst_code.append('        global _add_cell_right')
        self.lst_code.append('        _add_cell_right = self._add_cell_right')
        self.lst_code.append('        global _write_style')
        self.lst_code.append('        _write_style = self._write_style')
        text = '''
        for i, lst_row in enumerate(self.lst_cell):
            for r_cell in lst_row:
                self.r_row, self.r_col, self.border_idx, self.fill_idx, self.font_idx, self.alignment_idx, self.b_merged, self.b_range, \\
                self.row_span, self.col_span = r_cell
                if i == 0: # 第一行
                    col_name = get_column_letter(self.r_col)
                    self.w_sheet.column_dimensions[col_name].width = self.lst_col_width[self.r_col - 1] # 设置列宽
                
                self.w_row = self.r_row + self.dict_add_row.get(self.r_col, 0)
                self.w_col = self.r_col + self.dict_add_col.get(self.r_row, 0)
                self.w_cell = self.w_sheet.cell(row=self.w_row, column=self.w_col)
                
                if self.b_range: # 如果是合并就不处理
                    # self._copy_cell_style()
                    continue
                    
                elif self.b_merged: # 合并单元格起始位置
                    s_range = "%s%s:%s%s" % (get_column_letter(self.w_col), self.w_row, get_column_letter(self.w_col + self.col_span), 
                                             self.w_row + self.row_span)
                    self.w_sheet.merge_cells(s_range) # 合并单元格
                
                self._copy_cell_style()
                
                write_fun = getattr(self, "write_%s_%s" % (self.r_row, self.r_col), None)
                if write_fun is not None:
                    try:
                        write_fun()
                    except (BaseException,):
                        s_err_info = "%s%s内容错误，错误信息：%s" % (get_column_letter(self.r_col), self.r_row, traceback.format_exc())
                        raise ValueError(s_err_info)
                else:
                    # self.copy_cell_style(r_cell, self.w_sheet.cell(row=r_row, column=r_col), is_text=False)
                    pass
                    
            row_height = self.lst_row_height[self.r_row - 1] # 从0开始
            self.w_sheet.row_dimensions[self.w_row].height = row_height  # 设置行高
        self.w_book.save(self.output_file)

    def _w(self, text):
        if self.w_cell.value is None:
            self.w_cell.value = text
        else:
            self.w_cell.value = self.w_cell.value + text
'''
        self.lst_code.append(text)

        for node in self.lst_node:
            self._generate_code(node)
        self.lst_code.append('')
        return "\n".join(self.lst_code)


def generate_html(template_file,
                  compile=False,
                  max_row=200, max_col=100,
                  bg_color="#ffffff", font_color='#000000',
                  border_width=1, border_color='#000000',
                  font_size=14.67,
                  horizontal="center", vertical="center",
                  **kwargs):
    """
    compile:强制编译（excel转换成代码）
    col_width_percent：列宽使用百分比
    """
    compiler = PageCompiler(template_file=template_file, bg_color=bg_color, font_color=font_color,
                            border_width=border_width, border_color=border_color,
                            font_size=font_size, horizontal=horizontal, vertical=vertical)

    path, name = os.path.split(template_file)
    module_name = os.path.splitext(name)[0]
    if not os.path.exists("templates_py"):
        os.makedirs("templates_py")
        fh = open(os.path.join("templates_py", "__init__.py"), "w")
        fh.write("# !/usr/bin/python")
        fh.close()

    try:
        from xpinyin import Pinyin
        p = Pinyin()
        module_name_py = "%s_html" % p.get_pinyin(module_name, '').lower()
    except (BaseException,):
        module_name_py = "%s_html" % os.path.splitext(name)[0]

    file_path = os.path.join("templates_py", "%s.py" % module_name_py)
    # 如果文件存在，且excel修改时间小于
    if (not os.path.exists(file_path)) or compile or \
            (os.path.getmtime(file_path) < os.path.getmtime(template_file)):
        fh = codecs.open(file_path, "w", encoding='utf-8')
        result = compiler.compile_page(**kwargs)
        fh.write(result)
        fh.close()
        pass

    output = StringIO()
    module = importlib.import_module("templates_py.%s" % module_name_py)
    module.Page(output).create(**kwargs)
    return output


def generate_excel(template_file, output_file, compile=False, max_row=200, max_col=100, **kwargs):
    """
    根据模板文件生成excel文件
    :param template_file: 模板文件
    :param output_file: 生成文件
    :param compile: 是否强制编译
    :param max_row: 最大行
    :param max_col: 最大列
    :param kwargs: 参数
    """
    path, name = os.path.split(template_file)

    if not os.path.exists("templates_py"):
        os.makedirs("templates_py")
        fh = open(os.path.join("templates_py", "__init__.py"), "w")
        fh.write("# !/usr/bin/python")
        fh.close()

    module_name = os.path.splitext(name)[0]
    try:
        from xpinyin import Pinyin
        p = Pinyin()
        module_name_py = "%s_excel" % p.get_pinyin(module_name, '').lower()
    except (BaseException,):
        module_name_py = "%s_excel" % os.path.splitext(name)[0]

    file_path = os.path.join("templates_py", "%s.py" % module_name_py)

    # 如果文件存在，且excel修改时间小于
    if (not os.path.exists(file_path)) or compile or \
            (os.path.getmtime(file_path) < os.path.getmtime(template_file)):
        compiler = ExcelCompiler(template_file, 0, max_row, max_col)
        result = compiler.compile_page(**kwargs)
        fh = codecs.open(file_path, "w", encoding='utf-8')
        fh.write(result)
        fh.close()

    module = importlib.import_module("templates_py.%s" % module_name_py)
    module.Page(output_file).create(**kwargs)


if __name__ == "__main__":
    import time

    lst_progress = [
        ["第1周", "第2周", "第3周"],
        ["2018.03.19 -- 2018.03.25", "2018.04.16 -- 2018.04.22", "2018.03.19 -- 2018.03.25"],
        ["10%", "14%", "26%"],
        ["10%", "14%", "26%"],
    ]
    lst_test = [[1, 2, 3], [4, 5, 6]]
    week_date = "第2周"
    begin = time.time()
    # for i in range(300):
    #     output = generate_html(r"templates\20180412094402_16.xlsx", mandatory_compile=True, col_width_percent=True, _p={"aa": 2})
    #     fh = open("bb.html", "w")
    #     fh.write(output.getvalue())
    #     fh.close()
    # generate_excel("weizhi.xlsx", output_file="aa.xlsx",
    #                compile=True, max_row=40, max_col=30, aa=2,
    #                project_name="爱特his系统项目周报",
    #                begin_date="2018-03-20", end_date="2018-04-30", current_progress=53, next_progress=85,
    #                lst_progress=lst_progress, week_date=week_date, lst_test=lst_test)

    output = generate_html("weizhi.xlsx",
                           compile=True, aa=2,
                           project_name="爱特his系统项目周报",
                           begin_date="2018-03-20", end_date="2018-04-30", current_progress=53, next_progress=85,
                           lst_progress=lst_progress, week_date=week_date, lst_test=lst_test)

    fh = open("aa.html", "w")
    fh.write(output.getvalue())
    fh.close()
    #     break
    # end = time.time()
    # print(end - begin)
    # data = {}
    # lst_name = ["姓名1", "姓名2"]
    # lst_number = [100, 50]
    # lst_modify_number = [50, 25]
    # lst_data = [['10.1', '10.2'], [5, 4], [3, 2]]
    #
    # generate_excel(r"templates\bug_total.xlsx", "aa.xlsx", mandatory_compile=True,
    #                _p={"lst_name": lst_name, "lst_number": lst_number,
    #                    "lst_modify_number": lst_modify_number, "lst_data": lst_data})
    # w_book = openpyxl.Workbook()
    # w_sheet = w_book.active
    # w_sheet.cell(row=1, column=1).border = Border(left=Side(),right=Side(),top=Side(),bottom=Side())
    # w_sheet.cell(row=2, column=2).border = None
    # w_sheet.cell(row=1, column=1).fill = PatternFill(fgColor=Color(rgb='00000000',tint=0.0,type='rgb'),bgColor=Color(rgb='00000000',tint=0.0,type='rgb'))
    # w_sheet.cell(row=2, column=2).fill = PatternFill(patternType="solid",fgColor=Color(theme=0,tint=0.0,type='theme'))
    # w_book.save("cc.xlsx")
    # print(a)

    import cProfile

    # for i in range(300):
    #     generate_excel(r"templates\weekly.xlsx", "bb.xlsx", mandatory_compile=False, max_row=40, max_col=16,
    #                    project_name="爱特his系统项目周报",
    #                    begin_date="2018-03-20", end_date="2018-04-30", current_progress=53, next_progress=85,
    #                    lst_progress=lst_progress, week_date=week_date, lst_test=lst_test
    #                    )

    # 直接把分析结果打印到控制台
    # cProfile.run("test_generate_excel()")
    # test_generate_excel()
    end = time.time()
    print(end - begin)
