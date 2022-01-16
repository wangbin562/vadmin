import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from vadmin import widgets


class Compiler(object):
    def __init__(self, node, **kwargs):
        if isinstance(node, (tuple, list)):
            self.lst_node = node
        else:
            self.lst_node = [node, ]
        self.w_book = openpyxl.Workbook()
        self.w_sheet = self.w_book.active
        self.w_row = 1
        self.w_col = 1

    def write(self, lst_node):
        for i, node in enumerate(lst_node):
            if not node:
                continue

            if isinstance(node, str):
                self.write_text(node, self.w_row, self.w_col)
            else:
                widget_type = node.get("type", None)
                if widget_type == "text":
                    self.write_text(node, self.w_row, self.w_col)

                elif widget_type == "lite_table":
                    self.write_table(node)

                elif widget_type == "row":
                    self.w_row += 1
                    if node["height"]:
                        self.w_sheet.row_dimensions[self.w_row].height = node["height"]
                    self.w_col = 1

                elif widget_type == "col":
                    self.w_col += 1

    def write_text(self, node, row, col):
        cell = self.w_sheet.cell(row=row, column=col)
        if isinstance(node, (tuple, list)):
            cell.value = ""
            for n in node:
                if isinstance(n, widgets.Text):
                    cell.value += n.text
                elif isinstance(n, str):
                    cell.value += n
                elif n is None:
                    pass
                else:
                    cell.value += str(n)
        else:
            if isinstance(node, widgets.Text):
                cell.value = node.text
            elif node is None:
                pass
            else:
                cell.value = node

    def write_table(self, node):
        row = self.w_row
        col = self.w_col

        o_border = Border()
        if node.row_border:
            color = node.row_border["color"]
            o_border.top = Side(border_style='thin', color=color[1:])
            o_border.bottom = Side(border_style='thin', color=color[1:])

        if node.col_border:
            color = node.col_border["color"]
            o_border.left = Side(border_style='thin', color=color[1:])
            o_border.right = Side(border_style='thin', color=color[1:])

        for row_data in node.data:
            if node["row_height"]:
                self.w_sheet.row_dimensions[row].height = node["row_height"]
            elif node["min_row_height"]:
                self.w_sheet.row_dimensions[row].height = node["min_row_height"]

            col = self.w_col
            for cell_data in row_data:
                self.write_text(cell_data, row, col)
                cell = self.w_sheet.cell(row=row, column=col)

                o_alignment = Alignment(wrap_text=True)
                if node["horizontal"]:
                    o_alignment.horizontal = node["horizontal"]
                elif (col - 1) in node["col_horizontal"]:

                    o_alignment.horizontal = node["col_horizontal"][col - 1]

                if node["vertical"]:
                    if node["vertical"] == "middle":
                        o_alignment.vertical = "center"
                    else:
                        o_alignment.vertical = node["vertical"]

                elif (row - 1) in node["row_vertical"]:
                    if node["row_vertical"][row - 1] == "middle":
                        o_alignment.vertical = "center"
                    else:
                        o_alignment.vertical = node["row_vertical"][row - 1]

                cell.alignment = o_alignment
                cell.border = o_border

                col += 1
            row += 1

        self.w_row = row
        self.w_col = col

        for col, width in (node["col_width"] or {}).items():
            col_name = get_column_letter(col + 1)
            self.w_sheet.column_dimensions[col_name].width = round(width / 8.0, 2)

        for cells in node.get("merged_cells", []):
            begin, end = cells.split(":")
            begin_row, begin_col = begin.split("-")
            begin_row = int(begin_row)
            begin_col = int(begin_col)
            end_row, end_col = end.split("-")
            end_row = int(end_row)
            end_col = int(end_col)
            self.w_sheet.merge_cells(start_row=begin_row + 1, start_column=begin_col + 1,
                                     end_row=end_row + 1, end_column=end_col + 1)

        for cell, style in node.get("cell_style", {}).items():
            if "color" not in style.get("bg", {}):
                continue

            o_fill = PatternFill("solid", fgColor=style["bg"]["color"][1:])
            if ":" in cell:
                begin, end = cell.split(":")
                begin_row, begin_col = begin.split("-")
                begin_row = int(begin_row)
                begin_col = int(begin_col)
                end_row, end_col = end.split("-")
                end_row = int(end_row)
                end_col = int(end_col)
                for row in range(begin_row + 1, end_row + 2):
                    for col in range(begin_col + 1, end_col + 2):
                        cell = self.w_sheet.cell(row=row, column=col)
                        cell.fill = o_fill

            else:
                pass

    def save(self, path):
        self.w_book.save(path)


def generate(path, lst_node, **kwargs):
    """
    path:生成的文件路径
    lst_node:节点数据
    """
    obj = Compiler(lst_node, **kwargs)
    obj.write(obj.lst_node)
    obj.save(path)
