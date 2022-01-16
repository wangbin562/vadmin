# !/usr/bin/python
# -*- coding=utf-8 -*-
"""
v1.2
"""

import codecs
import collections
import copy
import importlib
import os
import time
import traceback

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from xpinyin import Pinyin

from utils import conver
from vadmin import widgets

BEGIN_KEY = "{{"
END_KEY = "}}"


def split_row_col(row_col):
    """
    分解行列
    :param row_col: 'B11' 'AB1' 'AB123'
    :return:
    """
    row = 0
    idx = 0
    for idx, key in enumerate(row_col):
        key_ascii = ord(key.upper())
        if (key_ascii >= 65) and (key_ascii <= 90):
            pass
        else:
            row = row_col[idx:]
            break

    col = column_index_from_string(row_col[0:idx])
    row = int(row)
    return col, row


def none_2_null_str(value):
    if value is None:
        return ""

    return value


def make_style(key, value, default, is_str=False, is_u=False):
    if isinstance(default, list):
        if value in default:
            return_value = ''
            return return_value

    if value == default:
        return_value = ''
        return return_value

    if is_str:
        if is_u:
            return_value = '%s=u"%s", ' % (key, value)
        else:
            return_value = '%s="%s", ' % (key, value)
    else:
        return_value = '%s=%s, ' % (key, value)

    return return_value


def make_color(color):
    if color is None:
        s_color = 'None'
    else:
        if color.type == "theme":
            s_color = "Color(theme=%s,tint=%s,type='%s')" % (color.theme, color.tint, color.type)
        elif color.type == "indexed":
            s_color = "Color(indexed=%s,tint=%s,type='%s')" % (color.indexed, color.tint, color.type)
        elif color.type == "rgb":
            s_color = "Color(rgb='%s',tint=%s,type='%s')" % (color.rgb, color.tint, color.type)
        else:
            s_color = "None"
    return s_color


def make_side(style, color):
    lst_none = [None, 'None']

    style = make_style('style', style, lst_none, True)
    color = make_style('color', color, lst_none)

    s_side = 'Side(%s%s' % (style, color)
    s_side = '%s)' % s_side.rstrip(",")
    return s_side


class BasePage(object):
    def __init__(self):
        self.r_cell = None  # 当前读的cell，list结构
        self.lst_border = []
        self.lst_fill = []
        self.lst_font = []
        self.lst_alignment = []

        self.dict_add_row = {}
        self.dict_add_col = {}
        self.dict_height = {}
        self.dict_width = {}
        self.r_row = None
        self.r_col = None
        self.w_row = None
        self.w_col = None
        self.w_cell = None

        self.lst_cell = []
        self.lst_row_height = []  # 模板表格的行高
        self.col_width = {}  # 模板表格的列宽
        self.o_widget = widgets.LiteTable(auto_wrap=True, scroll={"overflow": "drag"})
        self.o_widget.data = []
        self.o_widget.row_height = {}
        self.o_widget.col_width = {}
        self.o_widget.merged_cells = []

    def move_down(self, copy_style=True, copy_row_height=True, add_row=True):
        """
        向下增加一个单元格
        """
        if add_row:
            r_row, r_col, border, fill, font, alignment, b_merged, b_range, row_span, col_span = self.r_cell
            for col in range(r_col, r_col + col_span + 1):
                self.dict_add_row.setdefault(col, 0)
                self.dict_add_row[col] += (row_span + 1)
            self.w_row = r_row + self.dict_add_row.get(r_col, 0)
            self.w_col = r_col + self.dict_add_col.get(r_row, 0)
            if copy_row_height:
                for i in range(row_span + 1):
                    row = self.w_row + i
                    if row - 1 not in self.dict_height:
                        self.dict_height[row - 1] = r_row - 1

            self.expansion_cell()
            self.init_cell(copy_style=copy_style)

        else:
            self.w_row += 1
            self.expansion_cell()

    def move_right(self, copy_style=True, copy_col_width=True, add_col=False):
        """
        向右增加一个单元格
        """
        # if add_col:
        #     r_row, r_col, border, fill, font, alignment, b_merged, b_range, row_span, col_span = self.r_cell
        #     for row in range(r_row, r_row + row_span + 1):
        #         self.dict_add_col.setdefault(row, 0)
        #         self.dict_add_col[row] += (col_span + 1)
        #
        #     self.w_row = r_row + self.dict_add_row.get(r_col, 0)
        #     self.w_col = r_col + self.dict_add_col.get(r_row, 0)
        #     if copy_col_width:
        #         for i in range(col_span + 1):
        #             col = self.w_col + i
        #             if col - 1 not in self.dict_width:
        #                 self.dict_width[col - 1] = r_col - 1
        #
        #     self.expansion_cell()
        #     self.init_cell(copy_style=copy_style)
        # else:
        r_row, r_col, border, fill, font, alignment, b_merged, b_range, row_span, col_span = self.r_cell
        if copy_col_width:
            for i in range(col_span + 1):
                col = self.w_col + i
                if col not in self.dict_width:
                    self.col_width[col] = self.col_width[r_col - 1]

        self.w_col += 1
        self.expansion_cell()
        self.init_cell(copy_style=copy_style)

    def move(self, row=None, col=None):
        """移动到单元格"""
        r_row, r_col, border, fill, font, alignment, b_merged, b_range, row_span, col_span = self.r_cell

        if row is not None:
            r_row = row

        if col is not None:
            r_col = col

        self.r_cell = self.lst_cell[r_row - 1][r_col - 1]

    def merge_cells(self, begin_row, begin_col, end_row, end_col):
        """合并单元格"""
        self.o_widget.merged_cells.append("%s-%s:%s-%s" % (begin_row, begin_col,
                                                           end_row, end_col))

    def write_style(self, **kwargs):
        """
        写入样式
        """
        r_row, r_col, border_idx, fill_idx, font_idx, alignment_idx, b_merged, b_range, row_span, col_span = \
            self.r_cell
        o_panel = self.o_widget.data[r_row - 1][r_col - 1]

        if "background_color" in kwargs:
            o_panel.background_color = kwargs["background_color"]

        for o_widget in o_panel.get_widget():
            if isinstance(o_widget, widgets.Text):
                if "font_color" in kwargs:
                    o_widget.text_color = kwargs["font_color"]

                if "font_size" in kwargs:
                    o_widget.text_size = kwargs["font_size"]

        # if "wrap_text" in kwargs:
        #     o_cell.wrap_text = kwargs["wrap_text"]

    def write_widget(self, *args, **kwargs):
        r_row, r_col, border_idx, fill_idx, font_idx, alignment_idx, b_merged, b_range, row_span, col_span = \
            self.r_cell
        self.w_row = r_row + self.dict_add_row.get(r_col, 0)
        self.w_col = r_col + self.dict_add_col.get(r_row, 0)
        o_panel = self.o_widget.data[self.w_row - 1][self.w_col - 1]

        for param in args:
            if not isinstance(param, (tuple, list)):
                param = [param, ]

            for item in param:
                if isinstance(item, widgets.Widget):
                    o_panel.append(item)
                elif isinstance(item, (str, int, float, bool)):
                    self.write(item)
                else:
                    o_widget = widgets.Widget()
                    for k, v in item.items():
                        setattr(o_widget, k, v)
                    o_panel.append(o_widget)

    def write(self, text=None):
        if isinstance(text, (widgets.Widget,)):
            self.write_widget(text)
        elif text is not None:
            text = str(text)
            if text in ["\r", "\r\n", "\n"]:
                o_text = widgets.Row()
            else:
                o_text = widgets.Text(text=str(text))
                r_row, r_col, border_idx, fill_idx, font_idx, alignment_idx, b_merged, b_range, row_span, col_span = \
                    self.r_cell
                o_font = self.lst_font[font_idx]

                # font-size:40px
                if o_font.size:
                    o_text.font_size = o_font.size

                #  text-decoration: underline
                if o_font.underline:
                    o_text.font_decoration = "underline"

                if o_font.strike:
                    o_text.font_decoration = "line-through"

                if o_font.italic:
                    o_text.font_style = "italic"

                # 文字颜色
                color = conver.excel_color_2_rgb(o_font.color)
                if color and color not in ["#000000"]:
                    o_text.font_color = color

            # 写入样式
            o_panel = self.o_widget.data[self.w_row - 1][self.w_col - 1]
            o_panel.add_widget(o_text)

    def init_cell(self, copy_style=True, dict_style=None):
        r_row, r_col, border_idx, fill_idx, font_idx, alignment_idx, b_merged, b_range, row_span, col_span = self.r_cell
        # self.w_row = r_row + self.dict_add_row.get(r_col, 0)
        # self.w_col = r_col + self.dict_add_col.get(r_row, 0)
        if b_merged:
            self.o_widget.merged_cells.append("%s-%s:%s-%s" % (self.w_row - 1, self.w_col - 1,
                                                               self.w_row + row_span - 1,
                                                               self.w_col + col_span - 1))
        if not self.o_widget.data[self.w_row - 1][self.w_col - 1]:
            o_panel = widgets.Panel(width="100%", height="100%", padding=5,
                                    border={"width": [0, 0, 0, 0],
                                            "color": ["#000000", "#000000", "#000000", "#000000"]})
            if copy_style:
                # 边框, 默认有边框，要去除相邻单元格重复边框
                color = self.get_border_color("top")
                if color:
                    o_panel.border["width"][0] = 1
                    o_panel.border["color"][0] = color

                color = self.get_border_color("right")
                if color:
                    o_panel.border["width"][1] = 1
                    o_panel.border["color"][1] = color

                color = self.get_border_color("bottom")
                if color:
                    o_panel.border["width"][2] = 1
                    o_panel.border["color"][2] = color

                color = self.get_border_color("left")
                if color:
                    o_panel.border["width"][3] = 1
                    o_panel.border["color"][3] = color

                o_alignment = self.lst_alignment[alignment_idx]
                if o_alignment.horizontal:
                    o_panel.horizontal = o_alignment.horizontal

                if o_alignment.vertical:
                    o_panel.vertical = o_alignment.vertical
                else:
                    o_panel.vertical = "bottom"

                o_panel.background_color = conver.excel_color_2_rgb(self.lst_fill[fill_idx].fgColor)

            self.o_widget.data[self.w_row - 1][self.w_col - 1] = o_panel

    def expansion_cell(self):
        """分配数据"""
        if len(self.o_widget.data) < self.w_row:
            if self.lst_cell:
                self.o_widget.data.append([None] * len(self.lst_cell[0]))
            else:
                self.o_widget.data.append([None])

        row = self.o_widget.data[self.w_row - 1]
        row_len = len(row)
        for i in range(0, self.w_col):
            if i >= row_len:
                row.append(None)

    def get_border_color(self, direction):
        if direction == "top":
            r_row, r_col, border_idx, fill_idx, font_idx, alignment_idx, \
            b_merged, b_range, row_span, col_span = self.r_cell
            if self.w_row == 1:  # 非第一行肯定没有
                o_border = self.lst_border[border_idx]
                if o_border.top.style:
                    return conver.excel_color_2_rgb(o_border.top.color) or "#000000"

        elif direction == "right":
            r_row, r_col, border_idx, fill_idx, font_idx, alignment_idx, \
            b_merged, b_range, row_span, col_span = self.r_cell
            o_border = self.lst_border[border_idx]
            if o_border.right.style:
                return conver.excel_color_2_rgb(o_border.right.color) or "#000000"

            row = self.lst_cell[r_row - 1]
            if len(row) > (r_col + col_span):  # 下一列的左边
                o_border = self.lst_border[row[r_col + col_span - 1][2]]
                if o_border.right.style:
                    return conver.excel_color_2_rgb(o_border.right.color) or "#000000"

                o_border = self.lst_border[row[r_col + col_span][2]]
                if o_border.left.style:
                    return conver.excel_color_2_rgb(o_border.left.color) or "#000000"

        elif direction == "bottom":
            r_row, r_col, border_idx, fill_idx, font_idx, alignment_idx, \
            b_merged, b_range, row_span, col_span = self.r_cell
            o_border = self.lst_border[border_idx]
            if o_border.bottom.style:
                return conver.excel_color_2_rgb(o_border.bottom.color) or "#000000"

            if len(self.lst_cell) > (r_row + row_span):  # 下一行的头部
                o_border = self.lst_border[self.lst_cell[r_row + row_span - 1][r_col - 1][2]]
                if o_border.bottom.style:
                    return conver.excel_color_2_rgb(o_border.bottom.color) or "#000000"

                o_border = self.lst_border[self.lst_cell[r_row + row_span][r_col - 1][2]]
                if o_border.top.style:
                    return conver.excel_color_2_rgb(o_border.top.color) or "#000000"

        elif direction == "left":
            r_row, r_col, border_idx, fill_idx, font_idx, alignment_idx, \
            b_merged, b_range, row_span, col_span = self.r_cell
            if self.w_col == 1:
                o_border = self.lst_border[border_idx]
                if o_border.left.style:
                    return conver.excel_color_2_rgb(o_border.left.color) or "#000000"

        return None

    def create(self, **kwargs):
        self.init_para(**kwargs)
        for row, lst_row in enumerate(self.lst_cell):
            for col, r_cell in enumerate(lst_row):
                self.r_cell = r_cell
                r_row, r_col, border, fill, font, alignment, b_merged, b_range, row_span, col_span = r_cell

                if b_range:  # 合并的非第一个单元格
                    continue

                self.w_row = r_row + self.dict_add_row.get(r_col, 0)
                self.w_col = r_col + self.dict_add_col.get(r_row, 0)
                self.expansion_cell()
                self.init_cell()
                write_fun = getattr(self, "write_%s_%s" % (r_row, r_col), None)
                if write_fun is not None:
                    try:
                        write_fun()
                    except (BaseException,):
                        s_err_info = "%s%s单元格数据异常, %s" % (get_column_letter(r_col), r_row, traceback.format_exc())
                        print(s_err_info)
                        raise ValueError(s_err_info)

        lst_row = list(self.dict_height.keys())
        lst_row.sort()
        lst_height = []
        for row in lst_row:
            lst_height.append(self.lst_row_height[self.dict_height[row]])

        for i, row in enumerate(lst_row):
            height = lst_height[i]
            self.lst_row_height.insert(row, height)

        for row in range(len(self.o_widget.data)):
            self.o_widget.row_height[row] = self.lst_row_height[row]

        self.o_widget.col_width = self.col_width


class AbstractSyntaxNode:
    def __str__(self):
        return repr(self)

    def __repr__(self):
        return "<" + self.__class__.__name__ + ">"


class TextNode(AbstractSyntaxNode):
    """
    文本结点
    """

    def __init__(self, text, indent=0):
        self.text = text
        self.indent = indent

    def merge_text(self, text):
        self.text += text


class CommentNode(AbstractSyntaxNode):
    """
    注释结点
    """

    def __init__(self, text, indent=0):
        self.text = text
        self.indent = indent


class WrapNode(AbstractSyntaxNode):
    """
    换行结点
    """

    def __init__(self, indent=0):
        # self.text = text
        self.indent = indent
        pass


class ScriptNode(AbstractSyntaxNode):
    """
    脚本结点
    """

    def __init__(self, text, indent=0):
        self.text = text.rstrip()
        self.indent = indent


class Compiler(object):
    def __init__(self):
        self.lst_import = []
        self.lst_global = []
        self.lst_node = []
        self.def_num = 0
        self.block_name = None
        self.dict_method = collections.OrderedDict()
        self.work_book = None
        self.work_sheet = None
        self.current_idx = 0
        self.stream_len = None
        self.stream = None
        self.lst_cell = []
        self.lst_row_height = []
        self.col_width = {}
        self.max_row = 0
        self.max_col = 0
        self.lst_border = []
        self.lst_fill = []
        self.lst_font = []
        self.lst_alignment = []

    def process_instruction(self, instr, value, row, col):
        # 引入模块
        lst_node = []
        if instr == "import":
            value = "import %s" % value
            self.lst_import.append(value)

        elif instr == "from":
            value = "from %s " % value
            self.lst_import.append(value)

        else:
            print(instr)

        return lst_node

    def add_node(self, node):
        if self.block_name is not None:
            self.dict_method[self.block_name].append(node)
        else:
            self.lst_node.append(node)

    def calc_indent(self, key=None):
        if self.def_num > 0 and key == "def":
            indent = 8 * self.def_num

        elif self.def_num > 0:
            indent = 4 + 8 * self.def_num

        elif self.block_name and key == "def":
            indent = 4

        elif self.block_name:
            indent = 8
        else:
            indent = 8

        return indent

    def parse(self, row, col, row_span, col_span):
        """
        解释单元格
        """
        lst_node = []
        while True:
            begin_idx, end_idx = self.next_node()
            if (begin_idx < 0) or (end_idx < 0):
                if self.current_idx < self.stream_len:
                    begin_idx = self.stream_len
                else:
                    break

            if begin_idx != self.current_idx:
                text = self.stream[self.current_idx:begin_idx]
                # text = text.lstrip("\r\n")
                # text = text.rstrip()
                if text != "":
                    lst_text = text.split("\n")
                    if len(lst_text) == 1:
                        node = TextNode(text, self.calc_indent())
                        lst_node.append(node)
                    else:
                        for sig_text in lst_text:
                            node = TextNode(sig_text, self.calc_indent())
                            lst_node.append(node)
                            node = WrapNode(self.calc_indent())
                            lst_node.append(node)

                if end_idx < 0:
                    break

            key = self.stream[begin_idx + 2]
            if key == "#":
                text = self.stream[begin_idx + 3:end_idx]
                text = text.strip()
                lst_text = text.split(" ", 1)
                instr = lst_text[0].strip().lower()
                if len(lst_text) > 1:
                    # value = lst_text[1].strip().replace('"', '').replace("'", "")
                    value = lst_text[1].strip()
                else:
                    value = None
                lst_node.extend(self.process_instruction(instr, value, row, col))

            else:
                text = self.stream[begin_idx + 2:end_idx]
                text = text.lstrip("\r\n")
                text = text.rstrip("")
                if text.lower() == "end":
                    if self.def_num > 0:
                        self.def_num -= 1

                elif text[0:3].lower() == "def":
                    self.def_num += 1
                    lst_node.append(ScriptNode(text, self.calc_indent("def")))

                else:
                    lst_node.append(ScriptNode(text, self.calc_indent()))

            self.current_idx = end_idx + 2

        return lst_node

    def next_node(self):
        global BEGIN_KEY, END_KEY
        begin_idx = self.stream.find(BEGIN_KEY, self.current_idx)
        end_idx = self.stream.find(END_KEY, self.current_idx + 2)
        return begin_idx, end_idx

    def check_merged_cell(self, row, col):
        b_merged = False
        b_range = False
        row_span = 0
        col_span = 0
        row_col = '%s%s' % (get_column_letter(col), row)
        # for s_range in self.work_sheet.merged_cell_ranges:
        for s_range in self.work_sheet.merged_cells.ranges:
            s_range = str(s_range)
            begin_row_col, end_row_col = s_range.split(":")
            begin_col, begin_row = split_row_col(begin_row_col)
            end_col, end_row = split_row_col(end_row_col)

            if begin_row_col == row_col:  # 开始合并
                b_merged = True
                row_span = end_row - begin_row
                col_span = end_col - begin_col
                break

            elif (row >= begin_row) and (row <= end_row) and (col >= begin_col) and (col <= end_col):
                b_range = True
                break

        return b_merged, b_range, row_span, col_span

    def compile_page(self, file_name, sheet_idx=0, max_row=50, max_col=30, **kwargs):
        for key, val in kwargs.items():
            self.lst_global.append(key)
        self.max_row = max_row
        if isinstance(max_col, str):
            max_col = column_index_from_string(max_col)
        self.max_col = max_col
        self.work_book = openpyxl.load_workbook(file_name)
        self.work_sheet = self.work_book.worksheets[sheet_idx]
        self.parse_excel()
        return self.generate_code()

    def parse_excel(self):
        """
        解释文件
        """
        max_row = self.work_sheet.max_row
        max_col = self.work_sheet.max_column
        if max_row < self.max_row:
            self.max_row = max_row
        if max_col < self.max_col:
            self.max_col = max_col
        lst_row = [None] * self.max_col
        for i in range(self.max_row):
            self.lst_cell.append(copy.copy(lst_row))

        for row in range(1, self.max_row + 1):
            height = self.work_sheet.row_dimensions[row].height
            if height is None:
                height = 13.5

            height = conver.pt_2_px(height)
            self.lst_row_height.append(height)

            old_col_width = 9
            for col in range(1, self.max_col + 1):
                if row == 1:
                    col_name = get_column_letter(col)

                    if col_name in self.work_sheet.column_dimensions:
                        col_width = self.work_sheet.column_dimensions[col_name].width
                        if not self.work_sheet.column_dimensions[col_name].customWidth:
                            col_width = 9
                        old_col_width = col_width
                    else:
                        col_width = old_col_width
                    self.col_width[col - 1] = round(col_width * 8, 2)

                cell = self.work_sheet.cell(row=row, column=col)

                left_color = make_color(cell.border.left.color)
                right_color = make_color(cell.border.right.color)
                top_color = make_color(cell.border.top.color)
                bottom_color = make_color(cell.border.bottom.color)
                # diagonal_color = make_color(cell.border.diagonal.color)

                left_side = make_side(cell.border.left.style, left_color)
                right_side = make_side(cell.border.right.style, right_color)
                top_side = make_side(cell.border.top.style, top_color)
                bottom_side = make_side(cell.border.bottom.style, bottom_color)
                # diagonal_side = make_side(cell.border.diagonal.style, diagonal_color)
                # left = Side(), right = Side(), top = Side(),
                # bottom = Side(), diagonal = Side(), diagonal_direction = None,
                # vertical = None, horizontal = None, diagonalUp = False, diagonalDown = False,
                # outline = True, start = None, end = None
                # diagonal_direction = make_style('diagonal_direction', cell.border.diagonal_direction, None)
                # vertical = make_style('vertical', cell.border.vertical, None)
                # horizontal = make_style('horizontal', cell.border.horizontal, None)
                # diagonalUp = make_style('diagonalUp', cell.border.diagonalUp, False)
                # diagonalDown = make_style('diagonalDown', cell.border.diagonalDown, False)
                # outline = make_style('outline', cell.border.outline, True)
                # start = make_style('start', cell.border.start, None)
                # end = make_style('end', cell.border.end, None)
                border = "Border(left=%s,right=%s,top=%s,bottom=%s)" % (
                    left_side, right_side, top_side, bottom_side)
                if border not in self.lst_border:
                    self.lst_border.append(border)

                fgColor = make_color(cell.fill.fgColor)
                # bgColor = make_color(cell.fill.bgColor)
                patternType = make_style('patternType', cell.fill.patternType, None, True)
                fill = "PatternFill(%sfgColor=%s)" % (patternType, fgColor)
                if fill not in self.lst_fill:
                    self.lst_fill.append(fill)

                name = make_style('name', cell.font.name, None, True, True)
                # sz = make_style('sz', cell.font.sz, None)
                # b = make_style('b', cell.font.b, None)
                # i = make_style('i', cell.font.i, None)
                charset = make_style('charset', cell.font.charset, None)
                # u = make_style('u', cell.font.u, None)
                strike = make_style('strike', cell.font.strike, None)
                scheme = make_style('scheme', cell.font.scheme, None, True)
                family = make_style('family', cell.font.family, None)
                size = make_style('size', cell.font.size, None)
                bold = make_style('bold', cell.font.bold, None)
                italic = make_style('italic', cell.font.italic, None)
                underline = make_style('underline', cell.font.underline, None, is_str=True)
                vertAlign = make_style('vertAlign', cell.font.vertAlign, None)
                outline = make_style('outline', cell.font.outline, None)
                shadow = make_style('shadow', cell.font.shadow, None)
                condense = make_style('condense', cell.font.condense, None)
                extend = make_style('extend', cell.font.extend, None)
                color = make_color(cell.font.color)
                font = "Font(%s%s%s%s%s%s%s%s%s%s%s%s%s%scolor=%s)" % (
                    name, charset, strike, scheme, family, size, bold, italic, underline, vertAlign, outline,
                    shadow, condense, extend, color
                )
                if font not in self.lst_font:
                    self.lst_font.append(font)

                # horizontal = None, vertical = None,
                # textRotation = 0, wrapText = None, shrinkToFit = None, indent = 0, relativeIndent = 0,
                # justifyLastLine = None, readingOrder = 0, text_rotation = None,
                # wrap_text = None, shrink_to_fit = None, mergeCell = None
                horizontal = make_style('horizontal', cell.alignment.horizontal, None, True)
                vertical = make_style('vertical', cell.alignment.vertical, None, True)
                indent = make_style('indent', cell.alignment.indent, 0)
                relativeIndent = make_style('relativeIndent', cell.alignment.relativeIndent, 0)
                justifyLastLine = make_style('justifyLastLine', cell.alignment.justifyLastLine, None)
                readingOrder = make_style('readingOrder', cell.alignment.readingOrder, 0)
                text_rotation = make_style('text_rotation', cell.alignment.text_rotation, [None, 0])
                wrap_text = make_style('wrap_text', cell.alignment.wrap_text, None)
                shrink_to_fit = make_style('shrink_to_fit', cell.alignment.shrink_to_fit, None)
                alignment = "Alignment(%s%s%s%s%s%s%s%s%s)" % (
                    horizontal, vertical, indent, relativeIndent,
                    justifyLastLine, readingOrder, text_rotation, wrap_text, shrink_to_fit
                )
                if alignment not in self.lst_alignment:
                    self.lst_alignment.append(alignment)

                # 判读是否合并
                b_merged, b_range, row_span, col_span = self.check_merged_cell(row, col)
                self.lst_cell[row - 1][col - 1] = "%s, %s, %s, %s, %s, %s, %s, %s, %s, %s" % (
                    row, col, self.lst_border.index(border), self.lst_fill.index(fill),
                    self.lst_font.index(font), self.lst_alignment.index(alignment),
                    b_merged, b_range, row_span, col_span)
                if b_range:
                    continue

                self.stream = none_2_null_str(cell.value)
                if cell.data_type in ['n']:
                    self.stream = str(self.stream)
                self.stream_len = len(self.stream)
                self.current_idx = 0

                lst_node = self.parse(row, col, row_span, col_span)
                if lst_node:
                    # node = lst_node[0]
                    text = "def write_%s_%s(self):" % (row, col)
                    self.add_node(ScriptNode(text, 4))
                    # if isinstance(node, LoopRowNode) or isinstance(node, LoopColNode):
                    #     self.lst_node.extend(lst_node)
                    # else:
                    # text = "self.init_cell()"
                    # self.add_node(ScriptNode(text, 8))
                    self.lst_node.extend(lst_node)
                    self.add_node(ScriptNode("", 0))

    def _generate_code(self, node, lst_code):
        indent = " " * node.indent
        if isinstance(node, TextNode):
            lst_text = node.text.split("\n")
            count = len(lst_text) - 1
            for i, text in enumerate(lst_text):
                if i < count:
                    if "'" in text:
                        line = '%s_w("%s\\n")' % (indent, text)
                    else:
                        line = "%s_w('%s\\n')" % (indent, text)
                else:
                    if "'" in text:  # 最后一行不用换行
                        line = '%s_w("%s")' % (indent, text)
                    else:
                        line = "%s_w('%s')" % (indent, text)
                lst_code.append(line)

        elif isinstance(node, WrapNode):
            line = "%s_widget(widgets.Row(3))" % indent
            lst_code.append(line)

        # elif isinstance(node, AssignmentNode):
        #     line = "%s_w(str(%s))" % (indent, node.text)
        #     lst_code.append(line)

        elif isinstance(node, ScriptNode):
            for text in node.text.split("\n"):
                text = text.rstrip()
                text = text.replace("\t", "    ")
                lst_code.append(indent + text)

        elif isinstance(node, CommentNode):
            lst_code.append('')
            lst_code.append(indent + node.text)

        else:
            # raise CompilerError("invalid syntax element " + repr(c))
            print(node.text)

    def generate_code(self):
        """
        生成代码
        """
        lst_code = list()
        lst_code.append('# !/usr/bin/python')
        lst_code.append('# -*- coding: utf-8 -*-')
        lst_code.append('"""')
        lst_code.append('code generated by Epage compiler')
        lst_code.append('"""')
        lst_code.append('from openpyxl.styles import Border, Side, PatternFill, Font, Alignment')
        lst_code.append('from openpyxl.styles import Color')
        lst_code.append('from vadmin import widgets')
        lst_code.append('from utils.excel_2_vadmin import BasePage')
        lst_code.append('')

        if self.lst_import:
            for s_import in self.lst_import:
                lst_code.append(s_import)

        if self.lst_global:
            for s_global in self.lst_global:
                lst_code.append("%s = None" % s_global)

        lst_code.append('''_write_style = None
_widget = None
_move = None
_move_down = None
_move_right = None
_w = None


class Page(BasePage):
    def __init__(self):
        super().__init__()
        global _write_style
        _write_style = self.write_style
        global _widget
        _widget = self.write_widget
        global _move
        _move = self.move
        global _move_down
        _move_down = self.move_down
        global _move_right
        _move_right = self.move_right
        global _merge_cells
        _merge_cells = self.merge_cells
        global _w
        _w = self.write''')

        for board in self.lst_border:
            text = "        self.lst_border.append(%s)" % board
            lst_code.append(text)

        for fill in self.lst_fill:
            text = "        self.lst_fill.append(%s)" % fill
            lst_code.append(text)

        for font in self.lst_font:
            text = "        self.lst_font.append(%s)" % font
            lst_code.append(text)

        for alignment in self.lst_alignment:
            text = "        self.lst_alignment.append(%s)" % alignment
            lst_code.append(text)

        for i, lst_row in enumerate(self.lst_cell):
            text = "        self.lst_cell.append([])"
            lst_code.append(text)
            for cell in lst_row:
                text = "        self.lst_cell[%s].append([%s])" % (i, cell)
                lst_code.append(text)
        text = "        self.lst_row_height = %s" % str(self.lst_row_height)
        lst_code.append(text)
        text = "        self.col_width = %s" % str(self.col_width)
        lst_code.append(text)
        lst_code.append('')

        lst_code.append('    def init_para(self, **kwargs):')
        for s_global in self.lst_global:
            lst_code.append('        global %s' % s_global)
            lst_code.append('        %s = kwargs["%s"]' % (s_global, s_global))

        if not self.lst_global:
            lst_code.append('        pass')
        lst_code.append('')

        for node in self.lst_node:
            self._generate_code(node, lst_code)
        lst_code.append('')
        return "\n".join(lst_code)


def generate(template_file, is_compile=False, max_row=100, max_col=30,
             row_border=None, col_border=None, **kwargs):
    """
    template_file：模板文件
    is_compile:强制编译（excel转换成代码）
    max_row:最大行数
    max_col:最大列数
    """

    path, name = os.path.split(template_file)
    module_name = "excel_2_vadmin_%s" % os.path.splitext(name)[0]
    p = Pinyin()
    module_name_py = p.get_pinyin(module_name, '').lower()

    if not os.path.exists("templates_py"):
        os.makedirs("templates_py")
        open(os.path.join("templates_py", "__init__.py"), "w").write("# !/usr/bin/python")

    file_path = os.path.join("templates_py", "%s.py" % module_name_py)

    template_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(template_file)))
    template_py_time = None
    if os.path.exists(file_path):
        template_py_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(file_path)))

    # 如果文件存在，且excel修改时间小于
    if (not os.path.exists(file_path)) or is_compile or (template_py_time < template_time):
        compiler = Compiler()
        fh = codecs.open(file_path, "w", encoding='utf-8')
        result = compiler.compile_page(template_file, max_row=max_row, max_col=max_col, **kwargs)
        fh.write(result)
        fh.close()

    module = importlib.import_module("templates_py.%s" % module_name_py)
    importlib.reload(module)
    o_page = module.Page()
    o_page.o_widget.row_border = row_border
    o_page.o_widget.col_border = col_border
    o_page.create(**kwargs)
    o_page.o_widget.scroll = {"overflow": "auto"}
    o_page.o_widget.bg_color = "#FFFFFF"
    return o_page.o_widget


if __name__ == "__main__":
    pass
